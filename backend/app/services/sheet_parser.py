"""
Participant sheet parsing (spec 67).

Handles the shapes organisers actually send:

  Singles     Name | Club | City | Rating | Seed | Email | Phone
  Doubles     Team Name | Player 1 Name | Player 2 Name | Club | ...
  Roster      S.No | Name | Emp ID | Contact | Category | Partner Name | Partner Emp ID

The roster shape is the awkward one: one row per person, a free-text category
that often means "singles AND doubles", and reciprocal partner rows so each
team appears twice. It is handled by reading the whole sheet first, then
emitting entries -- a singles entry per person who entered singles, and one
entry per *unique* pair for doubles.

Identity comes from the employee id where the sheet has one. Names are
unreliable ("Srinivasan" is listed as partner "Srinivasan Big"), so an emp id
is turned into a deterministic email, which makes matching and re-import
dedupe exact without needing a schema change.
"""
import io

from typing import Any, Dict, List, Optional, Set, Tuple

# Values that mean "nothing here", not a real name.
BLANKS = {"", "na", "n/a", "n.a", "none", "nil", "-", "--", "nan", "null", "tbd"}


def _norm(value: Any) -> str:
    return str(value).strip().lower().replace("_", " ").replace(".", " ")


def _collapse(value: str) -> str:
    return "".join(ch for ch in value if ch.isalnum())


def pick_column(columns: List[str], candidates: List[str],
                exclude: Optional[List[str]] = None) -> Optional[str]:
    """Exact match beats prefix beats substring, so specific headers win."""
    excluded = set(x for x in (exclude or []) if x)
    available = [c for c in columns if c not in excluded]
    collapsed = {c: _collapse(c) for c in available}

    for stage in ("exact", "prefix", "contains"):
        for cand in candidates:
            target = _collapse(cand)
            for c in available:
                got = collapsed[c]
                if (stage == "exact" and got == target) \
                   or (stage == "prefix" and got.startswith(target)) \
                   or (stage == "contains" and target in got):
                    return c
    return None


TEAM_CANDIDATES = ["team name", "teamname", "team", "pair name", "pair", "doubles team"]
P1_CANDIDATES = ["player 1 name", "player1 name", "player 1", "player1", "p1 name", "p1",
                 "first player", "captain"]
P2_CANDIDATES = ["player 2 name", "player2 name", "player 2", "player2", "p2 name", "p2",
                 "partner name for doubles", "partner name", "partner", "second player"]
NAME_CANDIDATES = ["name", "player name", "full name", "participant", "player"]
TYPE_CANDIDATES = ["category", "type", "event", "categories"]
EMP_CANDIDATES = ["emp id", "empid", "employee id", "employee no", "emp no", "staff id", "id"]
PARTNER_EMP_CANDIDATES = ["partner emp id", "partner employee id", "partner id",
                          "player 2 emp id", "p2 emp id"]


def _text(value: Any) -> Optional[str]:
    """Cell text, or None when the cell is blank or a placeholder like 'NA'."""
    if value is None:
        return None
    # NaN, which openpyxl can hand back for an empty numeric cell. It is the one
    # float that is not equal to itself, and pd.isna is what this used to be.
    if isinstance(value, float) and value != value:
        return None
    text = str(value).strip()
    if text.lower() in BLANKS:
        return None
    return text


def _emp(value: Any) -> Optional[str]:
    """Employee id as a clean string; 2920908.0 from Excel becomes '2920908'."""
    text = _text(value)
    if text is None:
        return None
    try:
        return str(int(float(text)))
    except (TypeError, ValueError):
        return text


def parse_categories(raw: Optional[str], default_doubles: bool = False) -> Set[str]:
    """
    Read a free-text category cell into {'singles'} / {'doubles'} / both.

    Real sheets contain "Men's single and doubles", "Mens Single and doubles",
    "Men Singels and doubles" and "Men's Singles". Matching on the first letter
    (the previous approach) classified every one of those as neither, so they
    all fell back to the sheet-level default and the singles entries vanished.
    """
    if not raw:
        return {"doubles"} if default_doubles else {"singles"}

    text = raw.lower()
    # Tolerate the common misspellings of "singles" seen in real sheets.
    for wrong in ("singels", "singel", "sinlges", "signles"):
        text = text.replace(wrong, "singles")
    for wrong in ("doubls", "doubel", "dobules", "doblues"):
        text = text.replace(wrong, "doubles")

    found: Set[str] = set()
    if "single" in text:
        found.add("singles")
    if "double" in text or "doubles" in text:
        found.add("doubles")

    if not found:
        found = {"doubles"} if default_doubles else {"singles"}
    return found


def resolve_columns(columns: List[str]) -> Dict[str, Any]:
    team_col = pick_column(columns, TEAM_CANDIDATES)
    p1_col = pick_column(columns, P1_CANDIDATES, exclude=[team_col])
    p2_col = pick_column(columns, P2_CANDIDATES, exclude=[team_col, p1_col])
    used = [team_col, p1_col, p2_col]
    name_col = pick_column(columns, NAME_CANDIDATES, exclude=used)
    type_col = pick_column(columns, TYPE_CANDIDATES, exclude=used + [name_col])
    partner_emp_col = pick_column(columns, PARTNER_EMP_CANDIDATES)
    emp_col = pick_column(columns, EMP_CANDIDATES, exclude=used + [name_col, partner_emp_col])

    claimed = used + [name_col, type_col, emp_col, partner_emp_col]
    return {
        "team": team_col,
        "player1": p1_col,
        "player2": p2_col,
        "name": name_col,
        "type": type_col,
        "emp": emp_col,
        "partner_emp": partner_emp_col,
        "club": pick_column(columns, ["club", "academy", "association", "department"], exclude=claimed),
        "city": pick_column(columns, ["city", "town", "district", "location"], exclude=claimed),
        "rating": pick_column(columns, ["rating", "rank points", "points", "rate"], exclude=claimed),
        "seed": pick_column(columns, ["seed", "seeding"], exclude=claimed),
        "email": pick_column(columns, ["email", "e mail", "mail"], exclude=claimed),
        "phone": pick_column(columns, ["contact no", "contact", "phone", "mobile"], exclude=claimed),
        "partner_email": pick_column(columns, ["player 2 email", "partner email", "p2 email"]),
        "partner_phone": pick_column(columns, ["player 2 phone", "partner phone", "p2 phone"]),
    }


def synthetic_email(emp_id: str) -> str:
    """
    Stable address derived from an employee id.

    Gives every person a unique key when the sheet has no email column, so
    re-importing matches the same person instead of creating a duplicate.
    """
    safe = "".join(ch for ch in str(emp_id) if ch.isalnum()).lower()
    return f"emp{safe}@carromarena.com"


def _cell(row, col, default=None):
    if not col or col not in row.index:
        return default
    got = _text(row[col])
    return default if got is None else got


def _number(row, col, default, errors, label):
    raw = _cell(row, col)
    if raw is None:
        return default
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        errors.append(label)
        return default


class Row:
    """
    One sheet row, addressable by column name.

    Stands in for a pandas Series so parse_participants below needs no changes:
    it uses `row[column]` and `row.index`, and nothing else.
    """
    __slots__ = ("_values", "index")

    def __init__(self, values: Dict[str, Any], columns: List[str]):
        self._values = values
        self.index = columns

    def __getitem__(self, key: str) -> Any:
        return self._values.get(key)

    def __contains__(self, key: str) -> bool:
        return key in self._values


class Sheet:
    """
    A parsed spreadsheet: column names and rows.

    This replaced pandas. pandas was used for exactly three things -- read_csv,
    read_excel and isna -- and cost 60 MB itself, 31 MB of numpy, and from
    version 3.0 another 84 MB of pyarrow as a hard dependency. That took the
    serverless bundle to 229 MB against a 225 MB ceiling and stopped every
    deployment; it also cost 1.3 seconds on every cold start. openpyxl, already
    a dependency at 1.8 MB, reads xlsx, and the standard library reads csv.
    """
    __slots__ = ("_columns", "rows")

    def __init__(self, columns: List[str], rows: List[Dict[str, Any]]):
        self._columns = list(columns)
        self.rows = rows

    @property
    def columns(self) -> List[str]:
        return self._columns

    @columns.setter
    def columns(self, new_columns: List[str]) -> None:
        """
        Renaming the columns renames them on the rows too.

        parse_participants normalises the headers with `df.columns = [...]`, and
        a pandas DataFrame keys its rows BY the columns, so that one assignment
        moved both. Keeping the two independent looked harmless and made every
        lookup afterwards miss: the parser found the columns it wanted and then
        read None out of every cell.
        """
        new_columns = list(new_columns)
        if len(new_columns) == len(self._columns):
            mapping = dict(zip(self._columns, new_columns))
            self.rows = [{mapping.get(k, k): v for k, v in row.items()}
                         for row in self.rows]
        self._columns = new_columns

    def iterrows(self):
        for i, values in enumerate(self.rows):
            yield i, Row(values, self._columns)


def read_sheet(content: bytes, filename: str) -> Sheet:
    """Read an uploaded .csv, .xlsx or .xls into a Sheet."""
    name = (filename or "").lower()

    if name.endswith(".csv"):
        import csv
        # utf-8-sig so a spreadsheet exported from Excel does not put a BOM on
        # the first column name and make it unmatchable.
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        try:
            header = next(reader)
        except StopIteration:
            return Sheet([], [])
        columns = [str(c or "").strip() for c in header]
        rows = []
        for raw in reader:
            if not any((c or "").strip() for c in raw):
                continue
            rows.append({columns[i]: (raw[i] if i < len(raw) else None)
                         for i in range(len(columns))})
        return Sheet(columns, rows)

    from openpyxl import load_workbook
    # read_only for speed, data_only so a formula cell yields its value rather
    # than "=SUM(...)".
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return Sheet([], [])
    columns = [str(c).strip() if c is not None else "" for c in header]
    rows = []
    for raw in it:
        if raw is None or not any(c is not None and str(c).strip() for c in raw):
            continue
        rows.append({columns[i]: (raw[i] if i < len(raw) else None)
                     for i in range(len(columns))})
    wb.close()
    return Sheet(columns, rows)


def parse_participants(df: Sheet) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any]]:
    """
    Returns (entries, errors, meta).

    A person listed for both categories yields two entries: one singles, one
    doubles. Doubles entries are deduplicated by the unordered pair, so
    reciprocal rows produce a single team.
    """
    df.columns = [_norm(c) for c in df.columns]
    columns = list(df.columns)
    cols = resolve_columns(columns)
    errors: List[str] = []

    primary_col = cols["player1"] or cols["name"]
    if not primary_col:
        raise ValueError(
            "Could not find a player name column. Expected 'Name' (singles or roster sheets) "
            "or 'Player 1 Name' and 'Player 2 Name' (doubles sheets). Found: {}".format(
                ", ".join(columns) or "no columns")
        )

    sheet_is_doubles = bool(cols["player2"]) and not cols["type"]

    # ---- pass 1: read every row --------------------------------------------
    people: Dict[str, Dict[str, Any]] = {}   # key -> person
    rows: List[Dict[str, Any]] = []

    def key_for(emp: Optional[str], name: Optional[str]) -> str:
        return ("emp:" + emp) if emp else ("name:" + (name or "").lower())

    def remember(emp, name, club, city, rating, phone, email, own_row=False) -> str:
        k = key_for(emp, name)
        person = people.setdefault(k, {
            "key": k, "empId": emp, "name": name, "club": club, "city": city,
            "rating": rating, "phone": phone,
            "email": email or (synthetic_email(emp) if emp else None),
            "authoritative": False,
        })
        # A person's own row is the authority on their name; a partner reference
        # may spell it differently ("Srinivasan" listed as "Srinivasan Big").
        if own_row and name and not person.get("authoritative"):
            person["name"] = name
            person["authoritative"] = True
        # Otherwise only fill gaps from later, more complete mentions.
        for field, value in (("name", name), ("club", club), ("city", city),
                             ("phone", phone), ("email", email)):
            if value and not person.get(field):
                person[field] = value
        return k

    for idx, row in df.iterrows():
        line = idx + 2
        name = _cell(row, primary_col) or _cell(row, cols["name"])
        if not name:
            if any(_cell(row, c) for c in columns):
                errors.append(f"Row {line}: no player name, row skipped.")
            continue

        emp = _emp(row[cols["emp"]]) if cols["emp"] and cols["emp"] in row.index else None
        club = _cell(row, cols["club"], "Independent")
        city = _cell(row, cols["city"], "")
        rating = _number(row, cols["rating"], 1500, errors,
                         f"Row {line} ({name}): rating is not a number, defaulted to 1500.")
        seed = _number(row, cols["seed"], None, errors,
                       f"Row {line} ({name}): seed is not a number, ignored.")
        phone = _cell(row, cols["phone"])
        email = _cell(row, cols["email"])

        self_key = remember(emp, name, club, city, rating, phone, email, own_row=True)

        partner_name = _cell(row, cols["player2"])
        partner_emp = _emp(row[cols["partner_emp"]]) if cols["partner_emp"] and cols["partner_emp"] in row.index else None
        partner_key = None
        if partner_name or partner_emp:
            partner_key = remember(
                partner_emp, partner_name, club, city, rating,
                _cell(row, cols["partner_phone"]), _cell(row, cols["partner_email"]),
            )

        categories = parse_categories(_cell(row, cols["type"]), default_doubles=sheet_is_doubles)

        rows.append({
            "line": line, "selfKey": self_key, "partnerKey": partner_key,
            "categories": categories, "seed": seed,
            "teamName": _cell(row, cols["team"]),
        })

    # ---- pass 2: singles entries -------------------------------------------
    entries: List[Dict[str, Any]] = []
    singles_seen: Set[str] = set()

    def as_entry(person: Dict[str, Any], **extra) -> Dict[str, Any]:
        return {
            "name": person["name"], "empId": person.get("empId"),
            "club": person.get("club") or "Independent",
            "city": person.get("city") or None,
            "rating": person.get("rating") or 1500,
            "email": person.get("email"), "phone": person.get("phone"),
            "selected": True, **extra,
        }

    for r in rows:
        if "singles" not in r["categories"]:
            continue
        if r["selfKey"] in singles_seen:
            continue
        singles_seen.add(r["selfKey"])
        entries.append(as_entry(people[r["selfKey"]], type="singles", seed=r["seed"]))

    # ---- pass 3: doubles entries, one per unique pair ----------------------
    pairs_seen: Set[frozenset] = set()
    for r in rows:
        if "doubles" not in r["categories"]:
            continue
        if not r["partnerKey"]:
            errors.append(
                "Row {}: '{}' is entered for doubles but has no partner, so no team was formed.".format(
                    r["line"], people[r["selfKey"]]["name"])
            )
            continue
        if r["partnerKey"] == r["selfKey"]:
            errors.append(f"Row {r['line']}: a player cannot partner themselves.")
            continue

        pair = frozenset((r["selfKey"], r["partnerKey"]))
        if pair in pairs_seen:
            continue          # the reciprocal row for a team already seen
        pairs_seen.add(pair)

        me, partner = people[r["selfKey"]], people[r["partnerKey"]]
        entries.append(as_entry(me,
            type="doubles", seed=r["seed"],
            partnerName=partner["name"], partnerEmpId=partner.get("empId"),
            partnerEmail=partner.get("email"), partnerPhone=partner.get("phone"),
            teamName=r["teamName"] or f"{me['name']} & {partner['name']}",
        ))

    meta = {
        "detectedFormat": (
            "roster" if cols["type"] and cols["player2"]
            else "doubles" if sheet_is_doubles else "singles"
        ),
        "columnsDetected": {k: v for k, v in cols.items() if v},
        "peopleFound": len(people),
        "singlesCount": sum(1 for e in entries if e["type"] == "singles"),
        "doublesCount": sum(1 for e in entries if e["type"] == "doubles"),
    }
    return entries, errors, meta
